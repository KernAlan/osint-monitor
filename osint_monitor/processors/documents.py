"""Document intelligence: PDF extraction, table detection, and content reuse."""

from __future__ import annotations

import io
import logging
import re
import tempfile
from datetime import datetime
from typing import Optional

import numpy as np
import requests

from osint_monitor.collectors.base import BaseCollector
from osint_monitor.core.models import RawItemModel

logger = logging.getLogger(__name__)

MAX_PAGES = 50
MAX_TEXT_CHARS = 100_000


# ---------------------------------------------------------------------------
# PDF text extraction (multi-backend with fallback)
# ---------------------------------------------------------------------------

def _extract_with_pymupdf(pdf_path: str) -> dict:
    """Extract text using pymupdf (fitz) -- best quality."""
    import fitz  # pymupdf

    doc = fitz.open(pdf_path)
    pages = min(len(doc), MAX_PAGES)
    text_parts: list[str] = []
    for i in range(pages):
        text_parts.append(doc[i].get_text())

    metadata_raw = doc.metadata or {}
    metadata = {
        "title": metadata_raw.get("title", ""),
        "author": metadata_raw.get("author", ""),
        "creation_date": metadata_raw.get("creationDate", ""),
        "subject": metadata_raw.get("subject", ""),
        "producer": metadata_raw.get("producer", ""),
    }
    doc.close()

    full_text = "\n".join(text_parts)[:MAX_TEXT_CHARS]
    return {"text": full_text, "pages": pages, "metadata": metadata, "tables": []}


def _extract_with_pdfplumber(pdf_path: str) -> dict:
    """Extract text and tables using pdfplumber."""
    import pdfplumber

    tables: list[list[list[str]]] = []
    text_parts: list[str] = []

    with pdfplumber.open(pdf_path) as pdf:
        pages = min(len(pdf.pages), MAX_PAGES)
        for i in range(pages):
            page = pdf.pages[i]
            page_text = page.extract_text() or ""
            text_parts.append(page_text)

            page_tables = page.extract_tables() or []
            for tbl in page_tables:
                cleaned = [
                    [str(cell) if cell is not None else "" for cell in row]
                    for row in tbl
                ]
                tables.append(cleaned)

    full_text = "\n".join(text_parts)[:MAX_TEXT_CHARS]
    return {"text": full_text, "pages": pages, "metadata": {}, "tables": tables}


def _extract_with_pypdf2(pdf_path: str) -> dict:
    """Extract text using PyPDF2 -- basic fallback."""
    from PyPDF2 import PdfReader

    reader = PdfReader(pdf_path)
    pages = min(len(reader.pages), MAX_PAGES)
    text_parts: list[str] = []
    for i in range(pages):
        text_parts.append(reader.pages[i].extract_text() or "")

    metadata_raw = reader.metadata or {}
    metadata = {
        "title": str(getattr(metadata_raw, "title", "") or ""),
        "author": str(getattr(metadata_raw, "author", "") or ""),
        "creation_date": str(getattr(metadata_raw, "creation_date", "") or ""),
    }

    full_text = "\n".join(text_parts)[:MAX_TEXT_CHARS]
    return {"text": full_text, "pages": pages, "metadata": metadata, "tables": []}


def _download_to_temp(url: str) -> str:
    """Download a URL to a temporary file, return path."""
    resp = requests.get(url, timeout=30, stream=True)
    resp.raise_for_status()
    suffix = ".pdf"
    if ".docx" in url.lower():
        suffix = ".docx"
    elif ".xlsx" in url.lower():
        suffix = ".xlsx"
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=suffix)
    for chunk in resp.iter_content(chunk_size=65536):
        tmp.write(chunk)
    tmp.close()
    return tmp.name


def extract_pdf_text(pdf_path_or_url: str) -> dict:
    """Extract text from a PDF file path or URL.

    Tries pymupdf first, then pdfplumber, then PyPDF2.
    Downloads from URL to a temp file when needed.

    Returns:
        {"text": str, "pages": int, "metadata": dict, "tables": list}
    """
    import os

    # Download if URL
    is_url = pdf_path_or_url.startswith(("http://", "https://"))
    if is_url:
        try:
            pdf_path = _download_to_temp(pdf_path_or_url)
        except Exception as exc:
            logger.error("Failed to download PDF from %s: %s", pdf_path_or_url, exc)
            return {"text": "", "pages": 0, "metadata": {}, "tables": []}
    else:
        pdf_path = pdf_path_or_url

    try:
        # Try pymupdf first (best quality text extraction)
        try:
            result = _extract_with_pymupdf(pdf_path)
            logger.debug("Extracted PDF with pymupdf: %d pages", result["pages"])
            # Supplement with pdfplumber tables if pymupdf succeeded but no tables
            if not result["tables"]:
                try:
                    plumber_result = _extract_with_pdfplumber(pdf_path)
                    result["tables"] = plumber_result["tables"]
                except Exception:
                    pass
            return result
        except Exception as exc:
            logger.debug("pymupdf failed: %s, trying pdfplumber", exc)

        # Try pdfplumber
        try:
            result = _extract_with_pdfplumber(pdf_path)
            logger.debug("Extracted PDF with pdfplumber: %d pages", result["pages"])
            return result
        except Exception as exc:
            logger.debug("pdfplumber failed: %s, trying PyPDF2", exc)

        # Try PyPDF2
        try:
            result = _extract_with_pypdf2(pdf_path)
            logger.debug("Extracted PDF with PyPDF2: %d pages", result["pages"])
            return result
        except Exception as exc:
            logger.error("All PDF extractors failed for %s: %s", pdf_path_or_url, exc)
            return {"text": "", "pages": 0, "metadata": {}, "tables": []}
    finally:
        if is_url:
            try:
                os.unlink(pdf_path)
            except OSError:
                pass


# ---------------------------------------------------------------------------
# Heuristic table detection in plain text
# ---------------------------------------------------------------------------

def extract_tables_from_text(text: str) -> list[dict]:
    """Detect tables in plain text using heuristic patterns.

    Supports pipe-delimited, tab-delimited, and markdown tables.

    Returns:
        list of {"headers": list[str], "rows": list[list[str]], "format": str}
    """
    tables: list[dict] = []

    # Split into lines for scanning
    lines = text.split("\n")

    i = 0
    while i < len(lines):
        line = lines[i]

        # --- Markdown / pipe-delimited tables ---
        if "|" in line and line.strip().startswith("|"):
            table_lines: list[str] = []
            while i < len(lines) and "|" in lines[i] and lines[i].strip().startswith("|"):
                table_lines.append(lines[i])
                i += 1

            if len(table_lines) >= 2:
                parsed_rows: list[list[str]] = []
                separator_idx: int | None = None
                for idx, tl in enumerate(table_lines):
                    cells = [c.strip() for c in tl.strip().strip("|").split("|")]
                    # Detect separator row (e.g. |---|---|)
                    if all(re.match(r"^[-:]+$", c) for c in cells if c):
                        separator_idx = idx
                        continue
                    parsed_rows.append(cells)

                if parsed_rows:
                    headers = parsed_rows[0]
                    rows = parsed_rows[1:]
                    tables.append({
                        "headers": headers,
                        "rows": rows,
                        "format": "markdown",
                    })
            continue

        # --- Pipe-delimited (no leading pipe) ---
        if line.count("|") >= 2 and not line.strip().startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].count("|") >= 2:
                table_lines.append(lines[i])
                i += 1

            if len(table_lines) >= 2:
                parsed_rows = []
                for tl in table_lines:
                    cells = [c.strip() for c in tl.split("|")]
                    if all(re.match(r"^[-:]+$", c) for c in cells if c):
                        continue
                    parsed_rows.append(cells)

                if parsed_rows:
                    tables.append({
                        "headers": parsed_rows[0],
                        "rows": parsed_rows[1:],
                        "format": "pipe",
                    })
            continue

        # --- Tab-delimited tables ---
        if "\t" in line and line.count("\t") >= 1:
            table_lines = []
            col_count = line.count("\t") + 1
            while (
                i < len(lines)
                and "\t" in lines[i]
                and abs((lines[i].count("\t") + 1) - col_count) <= 1
            ):
                table_lines.append(lines[i])
                i += 1

            if len(table_lines) >= 2:
                parsed_rows = [
                    [c.strip() for c in tl.split("\t")]
                    for tl in table_lines
                ]
                tables.append({
                    "headers": parsed_rows[0],
                    "rows": parsed_rows[1:],
                    "format": "tab",
                })
            continue

        i += 1

    return tables


# ---------------------------------------------------------------------------
# Content reuse detection
# ---------------------------------------------------------------------------

def detect_content_reuse(
    session,
    text: str,
    threshold: float = 0.85,
) -> list[dict]:
    """Check if a document's content has been seen before.

    Splits text into paragraphs, embeds each, and compares against recent
    item embeddings stored in the database.

    Args:
        session: SQLAlchemy session.
        text: Document text to check.
        threshold: Cosine similarity threshold for a match.

    Returns:
        list of {"matching_item_id": int, "matching_title": str,
                 "similarity": float, "matching_paragraph": str}
    """
    from osint_monitor.processors.embeddings import (
        embed_texts,
        blob_to_embedding,
        cosine_similarity,
    )
    from osint_monitor.core.database import RawItem

    # Split into paragraphs (at least 60 chars to skip noise)
    paragraphs = [
        p.strip() for p in re.split(r"\n\s*\n", text) if len(p.strip()) >= 60
    ]
    if not paragraphs:
        return []

    # Embed paragraphs
    para_embeddings = embed_texts(paragraphs)  # (N, D)

    # Load recent items with embeddings (last 500)
    recent_items = (
        session.query(RawItem.id, RawItem.title, RawItem.embedding)
        .filter(RawItem.embedding.isnot(None))
        .order_by(RawItem.fetched_at.desc())
        .limit(500)
        .all()
    )

    if not recent_items:
        return []

    matches: list[dict] = []
    seen_item_ids: set[int] = set()

    for item_id, item_title, item_blob in recent_items:
        item_emb = blob_to_embedding(item_blob)

        for j, para_emb in enumerate(para_embeddings):
            sim = cosine_similarity(para_emb, item_emb)
            if sim >= threshold and item_id not in seen_item_ids:
                seen_item_ids.add(item_id)
                matches.append({
                    "matching_item_id": item_id,
                    "matching_title": item_title or "",
                    "similarity": round(float(sim), 4),
                    "matching_paragraph": paragraphs[j][:300],
                })

    # Sort by similarity descending
    matches.sort(key=lambda m: m["similarity"], reverse=True)
    return matches


# ---------------------------------------------------------------------------
# Generic document URL processing
# ---------------------------------------------------------------------------

def process_document_url(url: str, session=None) -> dict:
    """Process a document URL: detect type, extract content, check reuse.

    Args:
        url: URL of the document to process.
        session: Optional SQLAlchemy session for content reuse detection.

    Returns:
        {"url": str, "type": str, "text": str, "pages": int,
         "metadata": dict, "tables": list, "text_tables": list,
         "content_reuse": list}
    """
    # Detect type from URL or content-type header
    doc_type = "unknown"
    lower_url = url.lower()
    if lower_url.endswith(".pdf") or "/pdf" in lower_url:
        doc_type = "pdf"
    elif lower_url.endswith(".docx"):
        doc_type = "docx"
    elif lower_url.endswith(".xlsx"):
        doc_type = "xlsx"
    else:
        # Try HEAD request to detect content-type
        try:
            head = requests.head(url, timeout=10, allow_redirects=True)
            ct = head.headers.get("content-type", "").lower()
            if "pdf" in ct:
                doc_type = "pdf"
            elif "wordprocessingml" in ct or "msword" in ct:
                doc_type = "docx"
            elif "spreadsheetml" in ct or "ms-excel" in ct:
                doc_type = "xlsx"
        except Exception:
            pass

    result: dict = {
        "url": url,
        "type": doc_type,
        "text": "",
        "pages": 0,
        "metadata": {},
        "tables": [],
        "text_tables": [],
        "content_reuse": [],
    }

    if doc_type == "pdf":
        pdf_result = extract_pdf_text(url)
        result["text"] = pdf_result["text"]
        result["pages"] = pdf_result["pages"]
        result["metadata"] = pdf_result["metadata"]
        result["tables"] = pdf_result["tables"]
    elif doc_type in ("docx", "xlsx"):
        # Generic text extraction attempt
        try:
            import os

            tmp_path = _download_to_temp(url)
            try:
                if doc_type == "docx":
                    result["text"] = _extract_docx_text(tmp_path)
                elif doc_type == "xlsx":
                    result["text"] = _extract_xlsx_text(tmp_path)
            finally:
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass
        except Exception as exc:
            logger.warning("Failed to extract %s from %s: %s", doc_type, url, exc)
    else:
        # Attempt plain-text download
        try:
            resp = requests.get(url, timeout=30)
            resp.raise_for_status()
            result["text"] = resp.text[:MAX_TEXT_CHARS]
        except Exception as exc:
            logger.warning("Failed to download text from %s: %s", url, exc)

    # Detect tables in extracted text
    if result["text"]:
        result["text_tables"] = extract_tables_from_text(result["text"])

    # Content reuse detection
    if session and result["text"]:
        try:
            result["content_reuse"] = detect_content_reuse(session, result["text"])
        except Exception as exc:
            logger.warning("Content reuse detection failed: %s", exc)

    return result


def _extract_docx_text(path: str) -> str:
    """Extract plain text from a .docx file."""
    try:
        from docx import Document

        doc = Document(path)
        paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
        return "\n\n".join(paragraphs)[:MAX_TEXT_CHARS]
    except ImportError:
        logger.warning("python-docx not installed; cannot extract .docx")
        return ""


def _extract_xlsx_text(path: str) -> str:
    """Extract plain text from an .xlsx file."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        parts: list[str] = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            parts.append(f"=== Sheet: {sheet} ===")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                parts.append("\t".join(cells))
        wb.close()
        return "\n".join(parts)[:MAX_TEXT_CHARS]
    except ImportError:
        logger.warning("openpyxl not installed; cannot extract .xlsx")
        return ""


# ---------------------------------------------------------------------------
# DocumentCollector -- government document feeds
# ---------------------------------------------------------------------------

class DocumentCollector(BaseCollector):
    """Collector for government document feeds.

    Sources:
        - Federal Register API (free, no key required)
        - CRS Reports RSS from crsreports.congress.gov
    """

    FEDERAL_REGISTER_URL = "https://www.federalregister.gov/api/v1/documents.json"
    CRS_RSS_URL = "https://crsreports.congress.gov/rss/AllProducts"

    DOCUMENT_TYPES = (
        "presidential_documents",
        "rules",
        "proposed_rules",
        "notices",
    )

    def __init__(
        self,
        name: str = "government_documents",
        doc_types: list[str] | None = None,
        include_crs: bool = True,
        **kwargs,
    ):
        super().__init__(
            name=name,
            source_type="government_documents",
            url=self.FEDERAL_REGISTER_URL,
            **kwargs,
        )
        self.doc_types = doc_types or list(self.DOCUMENT_TYPES)
        self.include_crs = include_crs

    def collect(self) -> list[RawItemModel]:
        """Collect documents from Federal Register and optionally CRS Reports."""
        items: list[RawItemModel] = []
        items.extend(self._collect_federal_register())
        if self.include_crs:
            items.extend(self._collect_crs_reports())
        logger.info("[%s] Collected %d government documents", self.name, len(items))
        return items[: self.max_items]

    def _collect_federal_register(self) -> list[RawItemModel]:
        """Fetch recent documents from the Federal Register API."""
        items: list[RawItemModel] = []
        try:
            params = {
                "per_page": min(self.max_items, 20),
                "order": "newest",
                "conditions[type][]": self.doc_types,
            }
            resp = requests.get(
                self.FEDERAL_REGISTER_URL,
                params=params,
                timeout=30,
            )
            resp.raise_for_status()
            data = resp.json()

            for doc in data.get("results", []):
                pub_date = None
                if doc.get("publication_date"):
                    try:
                        pub_date = datetime.strptime(
                            doc["publication_date"], "%Y-%m-%d"
                        )
                    except ValueError:
                        pass

                pdf_url = doc.get("pdf_url", "")
                summary = doc.get("abstract") or doc.get("excerpt", "") or ""

                items.append(
                    RawItemModel(
                        title=doc.get("title", "Untitled Federal Register Document"),
                        content=summary[:5000],
                        url=doc.get("html_url", pdf_url),
                        published_at=pub_date,
                        source_name=self.name,
                        external_id=doc.get("document_number", ""),
                        fetched_at=datetime.utcnow(),
                    )
                )

            logger.debug(
                "Federal Register: fetched %d documents", len(items)
            )
        except Exception as exc:
            logger.error("Federal Register collection failed: %s", exc)
        return items

    def _collect_crs_reports(self) -> list[RawItemModel]:
        """Scrape CRS Reports from the RSS feed."""
        items: list[RawItemModel] = []
        try:
            import feedparser

            feed = feedparser.parse(self.CRS_RSS_URL)
            for entry in feed.entries[: self.max_items]:
                pub_date = None
                for attr in ("published_parsed", "updated_parsed"):
                    parsed = getattr(entry, attr, None)
                    if parsed:
                        try:
                            pub_date = datetime(*parsed[:6])
                        except Exception:
                            pass
                        break

                # Look for PDF link in enclosures or link
                pdf_url = ""
                for link in getattr(entry, "links", []):
                    if link.get("type", "") == "application/pdf" or link.get(
                        "href", ""
                    ).endswith(".pdf"):
                        pdf_url = link["href"]
                        break

                summary = re.sub(
                    r"<[^>]+>",
                    "",
                    getattr(entry, "description", "") or "",
                ).strip()

                items.append(
                    RawItemModel(
                        title=entry.get("title", "Untitled CRS Report"),
                        content=summary[:5000],
                        url=entry.get("link", pdf_url),
                        published_at=pub_date,
                        source_name=self.name,
                        external_id=entry.get("id") or entry.get("link", ""),
                        fetched_at=datetime.utcnow(),
                    )
                )

            logger.debug("CRS Reports: fetched %d reports", len(items))
        except ImportError:
            logger.warning("feedparser not installed; skipping CRS reports")
        except Exception as exc:
            logger.error("CRS Reports collection failed: %s", exc)
        return items
